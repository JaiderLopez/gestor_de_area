from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime
from io import BytesIO


def generate_template():
    """
    Genera un archivo Excel con plantilla para importar discos.
    Retorna un objeto BytesIO con el archivo Excel.
    """
    wb = Workbook()
    
    # Hoja 1: Disco
    ws_disco = wb.active
    ws_disco.title = "Disco"
    
    # Headers con estilo
    headers_disco = ['nombre', 'tipo', 'tamanio_gb', 'descripcion', 'estado']
    for col_num, header in enumerate(headers_disco, 1):
        cell = ws_disco.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="3699FF", end_color="3699FF", fill_type="solid")
    
    # Fila de ejemplo
    ws_disco.append(['Backup 2024', 'HDD', 500, 'Respaldo anual de contabilidad', 'BUENO'])
    
    # Hoja 2: Contenidos
    ws_contenidos = wb.create_sheet(title="Contenidos")
    
    headers_contenidos = ['disco_nombre', 'nombre', 'fecha_modificacion', 'peso_gb']
    for col_num, header in enumerate(headers_contenidos, 1):
        cell = ws_contenidos.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0BB783", end_color="0BB783", fill_type="solid")
    
    # Filas de ejemplo
    ws_contenidos.append(['Backup 2024', 'Facturas 2024', '2024-12-01', 15.5])
    ws_contenidos.append(['Backup 2024', 'Nomina', '2024-11-15', 8.2])
    
    # Ajustar ancho de columnas
    for ws in [ws_disco, ws_contenidos]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar en BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file


def parse_excel_import(file):
    """
    Lee un archivo Excel y extrae los datos de discos y contenidos.
    Retorna un diccionario con la estructura validada.
    """
    from openpyxl import load_workbook
    
    try:
        wb = load_workbook(file)
    except Exception as e:
        raise ValueError(f"Error al leer el archivo Excel: {str(e)}")
    
    # Verificar que existan las hojas requeridas
    if "Disco" not in wb.sheetnames:
        raise ValueError("El archivo debe contener una hoja llamada 'Disco'")
    if "Contenidos" not in wb.sheetnames:
        raise ValueError("El archivo debe contener una hoja llamada 'Contenidos'")
    
    ws_disco = wb["Disco"]
    ws_contenidos = wb["Contenidos"]
    
    # Parsear discos
    discos_data = []
    headers_disco = [cell.value for cell in ws_disco[1]]
    
    for row in ws_disco.iter_rows(min_row=2, values_only=True):
        if not any(row):  # Saltar filas vacías
            continue
        
        disco = dict(zip(headers_disco, row))
        # Normalizar el nombre del disco (trim whitespace)
        if disco.get('nombre'):
            disco['nombre'] = str(disco['nombre']).strip()
        disco['contenidos'] = []
        discos_data.append(disco)
    
    # Si no hay discos, retornar lista vacía
    if not discos_data:
        return []
    
    # Crear un mapa de nombres normalizados para búsqueda más eficiente
    # Clave: nombre normalizado (lowercase, sin espacios extra), Valor: disco original
    disco_map = {}
    for disco in discos_data:
        nombre_normalizado = disco.get('nombre', '').strip().lower()
        disco_map[nombre_normalizado] = disco
    
    print(f"DEBUG - Discos disponibles para asociar: {list(disco_map.keys())}")
    
    # Parsear contenidos y asociarlos a discos
    headers_contenidos = [cell.value for cell in ws_contenidos[1]]
    orphaned_contents = []  # Contenidos sin disco asociado
    
    for row_idx, row in enumerate(ws_contenidos.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        
        contenido = dict(zip(headers_contenidos, row))
        disco_nombre = contenido.pop('disco_nombre', None)
        
        if not disco_nombre:
            orphaned_contents.append({
                'row': row_idx,
                'reason': 'disco_nombre vacío',
                'data': contenido
            })
            continue
        
        # Normalizar el nombre del disco para búsqueda (trim + lowercase)
        disco_nombre_normalizado = str(disco_nombre).strip().lower()
        
        print(f"DEBUG - Buscando disco: '{disco_nombre}' (normalizado: '{disco_nombre_normalizado}')")
        
        # Buscar el disco usando el nombre normalizado
        disco_match = disco_map.get(disco_nombre_normalizado)
        
        if disco_match:
            # Convertir fecha a string si es datetime
            if isinstance(contenido.get('fecha_modificacion'), datetime):
                contenido['fecha_modificacion'] = contenido['fecha_modificacion'].strftime('%Y-%m-%d')
            disco_match['contenidos'].append(contenido)
            print(f"DEBUG - ✓ Contenido asociado al disco '{disco_match.get('nombre')}'")
        else:
            orphaned_contents.append({
                'row': row_idx,
                'reason': f'Disco "{disco_nombre}" no encontrado en la hoja Disco',
                'data': contenido
            })
            print(f"DEBUG - ✗ No se encontró disco para '{disco_nombre}'")
    
    # Si hay contenidos huérfanos, mostrar detalles
    if orphaned_contents:
        print(f"\n⚠️  Advertencia: {len(orphaned_contents)} contenido(s) no pudieron asociarse:")
        for orphan in orphaned_contents:
            print(f"  - Fila {orphan['row']}: {orphan['reason']}")
    
    return discos_data


def validate_import_data(discos_data):
    """
    Valida la estructura básica de los datos importados.
    Retorna lista de errores si los hay.
    """
    errors = []
    
    # Validar que haya al menos un disco
    if not discos_data or len(discos_data) == 0:
        errors.append({
            'row': 2,
            'sheet': 'Disco',
            'field': 'general',
            'message': 'No se encontraron discos para importar. La hoja "Disco" está vacía o solo contiene headers.'
        })
        return errors
    
    for idx, disco in enumerate(discos_data, start=2):  # start=2 porque row 1 es header
        # Validar campos requeridos
        if not disco.get('nombre'):
            errors.append({
                'row': idx,
                'sheet': 'Disco',
                'field': 'nombre',
                'message': 'El nombre es obligatorio'
            })
        
        if not disco.get('tipo'):
            errors.append({
                'row': idx,
                'sheet': 'Disco',
                'field': 'tipo',
                'message': 'El tipo es obligatorio'
            })
        elif disco['tipo'] not in ['HDD', 'SSD', 'CD/DVD', 'OTRO']:
            errors.append({
                'row': idx,
                'sheet': 'Disco',
                'field': 'tipo',
                'message': f"Tipo inválido. Debe ser: HDD, SSD, CD/DVD o OTRO"
            })
        
        if not disco.get('tamanio_gb'):
            errors.append({
                'row': idx,
                'sheet': 'Disco',
                'field': 'tamanio_gb',
                'message': 'El tamaño es obligatorio'
            })
        elif not isinstance(disco['tamanio_gb'], (int, float)) or disco['tamanio_gb'] <= 0:
            errors.append({
                'row': idx,
                'sheet': 'Disco',
                'field': 'tamanio_gb',
                'message': 'El tamaño debe ser un número positivo'
            })
        
        # Validar estado (opcional, con valor por defecto)
        if disco.get('estado'):
            if disco['estado'] not in ['BUENO', 'EN_RIESGO', 'DANADO']:
                errors.append({
                    'row': idx,
                    'sheet': 'Disco',
                    'field': 'estado',
                    'message': 'Estado inválido. Debe ser: BUENO, EN_RIESGO o DANADO'
                })
        else:
            # Si no viene en el Excel, asignar valor por defecto
            disco['estado'] = 'BUENO'
    
    return errors
