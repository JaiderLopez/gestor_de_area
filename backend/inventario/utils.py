from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime
from io import BytesIO

def generate_inventory_template():
    """
    Genera un archivo Excel con plantilla para importar dispositivos de inventario.
    Retorna un objeto BytesIO con el archivo Excel.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    
    # Headers con estilo
    headers = [
        'codigo_inventario', 'categoria', 'marca', 'modelo', 'serial', 
        'ubicacion', 'responsable', 'estado', 'fecha_compra', 
        'garantia_hasta', 'especificaciones'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="3699FF", end_color="3699FF", fill_type="solid")
    
    # Fila de ejemplo
    ws.append([
        'COMP-001', 'Portátil', 'Dell', 'Latitude 5420', '8H2K92', 
        'Oficina Gerencia', 'Juan Perez', 'ACTIVO', '2024-01-15', 
        '2027-01-15', 'i7 11th Gen, 16GB RAM, 512GB SSD'
    ])
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar en BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file

def parse_inventory_import(file):
    """
    Lee un archivo Excel y extrae los datos de inventario.
    Retorna una lista de diccionarios con la estructura validada.
    """
    try:
        wb = load_workbook(file)
    except Exception as e:
        raise ValueError(f"Error al leer el archivo Excel: {str(e)}")
    
    # Verificar que exista la hoja requerida
    sheet_name = "Inventario"
    if sheet_name not in wb.sheetnames:
        # Si no existe "Inventario", intentar usar la primera hoja
        ws = wb.active
    else:
        ws = wb[sheet_name]
    
    devices_data = []
    headers = [cell.value for cell in ws[1]]
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):  # Saltar filas vacías
            continue
        
        device = dict(zip(headers, row))
        
        # Normalizar strings básicos
        for key in ['codigo_inventario', 'marca', 'modelo', 'serial', 'ubicacion', 'responsable', 'categoria']:
            if device.get(key):
                device[key] = str(device[key]).strip()
        
        # Normalizar fechas
        for date_field in ['fecha_compra', 'garantia_hasta']:
            if device.get(date_field) and isinstance(device[date_field], datetime):
                device[date_field] = device[date_field].date()
        
        # Normalizar estado (upper case)
        if device.get('estado'):
            device['estado'] = str(device['estado']).strip().upper()
        else:
            device['estado'] = 'DISPONIBLE' # Default fallback
            
        devices_data.append(device)
    
    return devices_data

def validate_inventory_data(devices_data):
    """
    Valida la estructura básica de los datos importados.
    Retorna lista de errores si los hay.
    """
    errors = []
    
    if not devices_data:
        errors.append({
            'row': 2,
            'sheet': 'Inventario',
            'field': 'general',
            'message': 'No se encontraron datos para importar.'
        })
        return errors
        
    valid_states = ['ACTIVO', 'DISPONIBLE', 'EN_REPARACION', 'DAÑADO', 'BAJA']
    
    for idx, device in enumerate(devices_data, start=2):
        # Validar campos obligatorios
        if not device.get('codigo_inventario'):
            errors.append({
                'row': idx,
                'sheet': 'Inventario',
                'field': 'codigo_inventario',
                'message': 'El código de inventario es obligatorio'
            })
            
        if not device.get('categoria'):
             errors.append({
                'row': idx,
                'sheet': 'Inventario',
                'field': 'categoria',
                'message': 'La categoría es obligatoria'
            })
            
        if not device.get('marca'):
            errors.append({
                'row': idx,
                'sheet': 'Inventario',
                'field': 'marca',
                'message': 'La marca es obligatoria'
            })
            
        if not device.get('modelo'):
            errors.append({
                'row': idx,
                'sheet': 'Inventario',
                'field': 'modelo',
                'message': 'El modelo es obligatorio'
            })
            
        # Validar estado
        if device.get('estado') and device['estado'] not in valid_states:
            errors.append({
                'row': idx,
                'sheet': 'Inventario',
                'field': 'estado',
                'message': f'Estado inválido. Debe ser: {", ".join(valid_states)}'
            })
            
    return errors
